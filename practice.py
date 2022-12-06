import spacy

# nlp = spacy.load("en_core_web_md")

# doc1 = nlp("I like salty fries & hamburgers.")
# doc2 = nlp("Fast food tastes very good.")
#
# print(doc1, "<->", doc2, doc1.similarity(doc2))

# ---------------------------------- Entity Ruler--------------------------------------------------
# 1) Rule based approach
# 2) Machine learning approach
from spacy.tokens import Span

# nlp = spacy.load("en_core_web_sm")
# text = "West Chestertenfieldville was referenced as Mr. Deeds"
#
# patterns = [
#     {"labels": "GPE", "pattern": "West Chestertenfieldville"},
#     {"labels": "FILM", "pattern": "Mr. Deeds"}
# ]
#
# ruler = nlp.add_pipe("entity_ruler", before="ner")
# ruler.add_patterns(patterns)
#
# doc = nlp(text)
#
# for ent in doc.ents:
#     print(ent.text, ent.label_)
#---------------------------------------Rule-based -------------------------------------------------------

# from spacy.matcher import Matcher
#
# AA: list[str] = ["get", "place", "approx"]
# AB = ["get", "place", "exact"]
# AC = ["get", "place", "approx", "bulky"]
# AD = ["get", "place", "exact", "bulky"]
#
from spacy.tokens import Span

# nlp = spacy.load("en_core_web_sm")
# matcher = Matcher(nlp.vocab)
#
# pattern1 = [{'LOWER': 'get'}]
# pattern2 = [{'LOWER': 'place'}]
# pattern3 = [{'LOWER': 'approx'}]
# pattern4 = [{'LOWER': 'exact'}]
# pattern5 = [{'LOWER': 'bulky'}]
#
# matcher.add("AA", [pattern1, pattern2, pattern3])
# #matcher.add("AB", [pattern1, pattern2, pattern4])
#
# doc = nlp("Getting pin and placing on screw approximately")
# text = "Get pin and place on screw approx"
# matches = matcher(doc)
# #print((matches))
# AA_bin =False
# AB_bin = False
#
# for match_id, start, end in matches:
#     string_id = nlp.vocab.strings[match_id]  # Get string representation]
#     #print(string_id)
#     span = doc[start:end]  # The matched span
#     #print(match_id, string_id, start, end, span.text)
#
# for j in matches:
#     for i in j:
#         #print(i)
#--------------------------------------------------------------------------------------------------------------------

# import spacy
# from spacy.matcher import PhraseMatcher
#
# nlp = spacy.load("en_core_web_sm")
# phrase_matcher = PhraseMatcher(nlp.vocab)
#
# AA_words = [nlp.make_doc(text) for text in ["get", "place", "approx"]]
# #AB_words = [nlp.make_doc(text) for text in ["get", "place", "exact"]]
#
# phrase_matcher.add('AA', None, *AA_words)
# #phrase_matcher.add('AB', None, *AB_words)

# text = 'Get and place screw approx.'
# #         'julie had a little goat',
# 'julie enjoys eating pizza', 'mary went to the market',
# 'in the market there was a lamb', 'my goat likes to drink coffee',
# 'tara throws a ball for her goat', 'a goat and a kangaroo can often be friends',
# 'tara and mary like to drink beer']

# doc = nlp(text)
# matches = phrase_matcher(doc)
# for match_id, start, end in matches:
#         rule_id = nlp.vocab.strings[match_id]  # get the unicode ID, i.e. 'CategoryID'
#         span = doc[start : end]  # get the matched slice of the doc
#         print(rule_id, span.text)











# for fragment in text:
#     doc = nlp(fragment)
#     matches = phrase_matcher(doc)
#     print(matches[0])
#     rule_ids = {nlp.vocab.strings[match[0]] for match in matches}
#     print(rule_ids)
#     # if {'AA'}.IN(rule_ids):







   #     print('AA')
#     # if {'AB'}.IN(rule_ids):
#     #     print('AB')

# --------------- final code.......................................................................................................

# import xlwings as xw
import random
from typing import List, Any
import pandas as pd
from openpyxl.reader.excel import load_workbook
import spacy

# value1 = []
# code_df: list[Any]


# @xw.sub
# def main(x):
#     wb = xw.Book.caller()
#     wb.sheets(0).range("A1").value = "Hello wlwings!"
#     nlp = spacy.load("en_core_web_sm")
#     doc = nlp(x)
#     sentence_lemma = []
#     for token in doc:
#         sentence_lemma.append(token.lemma_)
#
#     value = mtm_code(sentence_lemma)
#     # value1.append(value)
#     return value
#
#
#
#
# def mtm_code(decription):
#     aa_words = ['get', 'place', 'approx']
#     ab_words = ["get", "place", "exact"]
#     ac_words = ["get", "place", "approx", "bulky"]
#     ad_words = ["get", "place", "exact", "bulky"]
#     pa_words = ["place", "approx"]
#     pb_words = ["place", "exact"]
#     ha_words = ["handle", "approx"]
#     hb_words = ["handle"]
#     ba_words = ["operate"]
#     bb_words = ["operate", "compound"]
#     za_words = ["motion", "without shift", "<=10"]
#     zb_words = ["motion", "without shift", ">10 - 30"]
#     zc_words = ["motion", "without shift", ">30 - 80"]
#     zd_words = ["motion", "with shift", "<=20"]
#     ze_words = ["motion", "with shift", "> 20 - 45"]
#     zf_words = ["motion", "with shift", "> 45 - 100"]
#     zz_words = ["tighten"]
#     zz1_words = ["loosen"]
#     ka_words = ["walk"]
#     kb_words = ["bend"]
#     kc_words = ["sit"]
#     kc1_words = ["stand"]
#     va_words = ["visual"]
#
#     if set(aa_words).issubset(decription):
#         return 'AA'
#     if set(ab_words).issubset(decription):
#         return 'AB'
#     if set(ac_words).issubset(decription):
#         return 'AC'
#     if set(ad_words).issubset(decription):
#         return 'AD'
#     if set(pa_words).issubset(decription):
#         return 'PA'
#     if set(pb_words).issubset(decription):
#         return 'PB'
#     if set(ha_words).issubset(decription):
#         return 'HA'
#     if set(hb_words).issubset(decription):
#         return 'HB'
#     if set(ba_words).issubset(decription):
#         return 'BA'
#     if set(bb_words).issubset(decription):
#         return 'BB'
#     if set(za_words).issubset(decription):
#         return 'ZA'
#     if set(zb_words).issubset(decription):
#         return 'ZB'
#     if set(zc_words).issubset(decription):
#         return 'ZC'
#     if set(zd_words).issubset(decription):
#         return 'ZD'
#     if set(ze_words).issubset(decription):
#         return 'ZE'
#     if set(zf_words).issubset(decription):
#         return 'ZF'
#     if set(zz_words).issubset(decription):
#         return 'ZZ'
#     if set(zz1_words).issubset(decription):
#         return 'ZZ'
#     if set(ka_words).issubset(decription):
#         return 'KA'
#     if set(kb_words).issubset(decription):
#         return 'KB'
#     if set(kc_words).issubset(decription):
#         return 'KC'
#     if set(kc1_words).issubset(decription):
#         return 'KC'
#     if set(va_words).issubset(decription):
#         return 'VA'

# wb = load_workbook(filename='MTM_MEK_Template.xlsm', read_only=False, keep_vba=True)
# ws = wb["SOS_ERGO - AS-IS"]
# df = pd.DataFrame(ws.values)
# print(df)

# print(df.iloc[0])
# print(df[9].values[1:])
#     for row in df[9].values[1:]:
#         # print(row)
#         if row is not None:

# code_df.insert(0, 0)
#
# print(code_df)
# # df['Code'] = code_df
# # print(df)
# df['Code'] = code_df
# # ws["SOS_ERGO - AS-IS"] = df
# #
# print(df['Code'])

# wb.save('MTM_MEK_Template.xlsm')








#
# import xlwings as xw
# import random
# from typing import List, Any
# import pandas as pd
# from openpyxl.reader.excel import load_workbook
# import spacy
#
# # value1 = []
# # code_df: list[Any]
#
#
# @xw.sub
# def main(x):
#     wb = xw.Book.caller()
#     wb.sheets(0).range("A1").value = "Hello wlwings!"
#     nlp = spacy.load("en_core_web_sm")
#     doc = nlp(x)
#     sentence_lemma = []
#     for token in doc:
#         sentence_lemma.append(token.lemma_)
#
#     value = mtm_code(sentence_lemma)
#     # value1.append(value)
#     return value
#



# def mtm_code(decription):
#     aa_words = ['get', 'place', 'approx']
#     ab_words = ["get", "place", "exact"]
#     ac_words = ["get", "place", "approx", "bulky"]
#     ad_words = ["get", "place", "exact", "bulky"]
#     pa_words = ["place", "approx"]
#     pb_words = ["place", "exact"]
#     ha_words = ["handle", "approx"]
#     hb_words = ["handle"]
#     ba_words = ["operate"]
#     bb_words = ["operate", "compound"]
#     za_words = ["motion", "without shift", "<=10"]
#     zb_words = ["motion", "without shift", ">10 - 30"]
#     zc_words = ["motion", "without shift", ">30 - 80"]
#     zd_words = ["motion", "with shift", "<=20"]
#     ze_words = ["motion", "with shift", "> 20 - 45"]
#     zf_words = ["motion", "with shift", "> 45 - 100"]
#     zz_words = ["tighten"]
#     zz1_words = ["loosen"]
#     ka_words = ["walk"]
#     kb_words = ["bend"]
#     kc_words = ["sit"]
#     kc1_words = ["stand"]
#     va_words = ["visual"]
#
#     if set(aa_words).issubset(decription):
#         return 'AA'
#     if set(ab_words).issubset(decription):
#         return 'AB'
#     if set(ac_words).issubset(decription):
#         return 'AC'
#     if set(ad_words).issubset(decription):
#         return 'AD'
#     if set(pa_words).issubset(decription):
#         return 'PA'
#     if set(pb_words).issubset(decription):
#         return 'PB'
#     if set(ha_words).issubset(decription):
#         return 'HA'
#     if set(hb_words).issubset(decription):
#         return 'HB'
#     if set(ba_words).issubset(decription):
#         return 'BA'
#     if set(bb_words).issubset(decription):
#         return 'BB'
#     if set(za_words).issubset(decription):
#         return 'ZA'
#     if set(zb_words).issubset(decription):
#         return 'ZB'
#     if set(zc_words).issubset(decription):
#         return 'ZC'
#     if set(zd_words).issubset(decription):
#         return 'ZD'
#     if set(ze_words).issubset(decription):
#         return 'ZE'
#     if set(zf_words).issubset(decription):
#         return 'ZF'
#     if set(zz_words).issubset(decription):
#         return 'ZZ'
#     if set(zz1_words).issubset(decription):
#         return 'ZZ'
#     if set(ka_words).issubset(decription):
#         return 'KA'
#     if set(kb_words).issubset(decription):
#         return 'KB'
#     if set(kc_words).issubset(decription):
#         return 'KC'
#     if set(kc1_words).issubset(decription):
#         return 'KC'
#     if set(va_words).issubset(decription):
#         return 'VA'
#
# wb = load_workbook(filename='MTM_MEK_Template.xlsm', read_only=False, keep_vba=True)
# ws = wb["SOS_ERGO - AS-IS"]
# df = pd.DataFrame(ws.values)
# print(df)
#
# # print(df.iloc[0])
# # print(df[9].values[1:])
#     for row in df[9].values[1:]:
#         # print(row)
#         if row is not None:
#
# code_df.insert(0, 0)

# print(code_df)
# # df['Code'] = code_df
# # print(df)
# df['Code'] = code_df
# # ws["SOS_ERGO - AS-IS"] = df
# #
# print(df['Code'])

# wb.save('MTM_MEK_Template.xlsm')
